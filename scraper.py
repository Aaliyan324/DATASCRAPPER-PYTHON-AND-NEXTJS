import re
import time
import requests
import pandas as pd
from datetime import datetime
import os
from fpdf import FPDF
from fpdf.enums import XPos, YPos

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# CONFIGURATION
# ============================================================

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OVERPASS_URL = "https://overpass-api.de/api/interpreter"

HEADERS = {
    "User-Agent": "FreeBusinessDataScraper/1.0"
}


# ============================================================
# BUSINESS TYPES (expanded)
# ============================================================

BUSINESS_TYPES = {

    # Restaurants & Food
    "restaurant": ("amenity", "restaurant"),
    "restaurants": ("amenity", "restaurant"),
    "resturant": ("amenity", "restaurant"),
    "resturants": ("amenity", "restaurant"),
    "eatery": ("amenity", "restaurant"),
    "food": ("amenity", "restaurant"),
    "fast food": ("amenity", "fast_food"),
    "fastfood": ("amenity", "fast_food"),
    "pizza": ("amenity", "fast_food"),
    "diner": ("amenity", "restaurant"),
    "cafe": ("amenity", "cafe"),
    "cafes": ("amenity", "cafe"),
    "coffee shop": ("amenity", "cafe"),
    "coffee": ("amenity", "cafe"),
    "bakery": ("shop", "bakery"),
    "bakeries": ("shop", "bakery"),

    # Hotels & Accommodation
    "hotel": ("tourism", "hotel"),
    "hotels": ("tourism", "hotel"),
    "lodging": ("tourism", "hotel"),
    "inn": ("tourism", "hotel"),
    "motel": ("tourism", "motel"),
    "guest house": ("tourism", "guest_house"),
    "guesthouse": ("tourism", "guest_house"),

    # Shopping
    "supermarket": ("shop", "supermarket"),
    "supermarkets": ("shop", "supermarket"),
    "grocery": ("shop", "supermarket"),
    "store": ("shop", "supermarket"),
    "clothes": ("shop", "clothes"),
    "clothing": ("shop", "clothes"),
    "apparel": ("shop", "clothes"),
    "fashion": ("shop", "clothes"),
    "electronics": ("shop", "electronics"),
    "mobile": ("shop", "mobile_phone"),
    "furniture": ("shop", "furniture"),
    "bookstore": ("shop", "books"),
    "books": ("shop", "books"),

    # Health & Medical
    "hospital": ("amenity", "hospital"),
    "hospitals": ("amenity", "hospital"),
    "clinic": ("amenity", "clinic"),
    "clinics": ("amenity", "clinic"),
    "pharmacy": ("amenity", "pharmacy"),
    "pharmacies": ("amenity", "pharmacy"),
    "chemist": ("amenity", "pharmacy"),
    "drugstore": ("amenity", "pharmacy"),
    "dentist": ("amenity", "dentist"),
    "dentists": ("amenity", "dentist"),
    "doctor": ("amenity", "doctors"),
    "medical": ("amenity", "doctors"),

    # Education
    "school": ("amenity", "school"),
    "schools": ("amenity", "school"),
    "college": ("amenity", "college"),
    "university": ("amenity", "university"),
    "institute": ("amenity", "college"),

    # Banking & Finance
    "bank": ("amenity", "bank"),
    "banks": ("amenity", "bank"),
    "atm": ("amenity", "atm"),
    "finance": ("amenity", "bank"),

    # Fitness & Leisure
    "gym": ("leisure", "fitness_centre"),
    "gyms": ("leisure", "fitness_centre"),
    "fitness": ("leisure", "fitness_centre"),
    "park": ("leisure", "park"),
    "playground": ("leisure", "playground"),

    # Personal Care
    "salon": ("shop", "hairdresser"),
    "salons": ("shop", "hairdresser"),
    "barber": ("shop", "hairdresser"),
    "beauty": ("shop", "beauty"),
    "spa": ("shop", "beauty"),

    # Places of Worship
    "mosque": ("amenity", "mosque"),
    "church": ("amenity", "church"),
    "temple": ("amenity", "temple"),

    # Other
    "post office": ("amenity", "post_office"),
    "postoffice": ("amenity", "post_office"),
    "police": ("amenity", "police"),
    "library": ("amenity", "library"),
}


# ============================================================
# ROMAN URDU COMMAND WORDS (expanded)
# ============================================================

REMOVE_WORDS = [

    # English
    "find",
    "search",
    "show",
    "get",
    "give",
    "need",
    "want",
    "looking",
    "looking for",
    "list",
    "near",
    "around",
    "in",
    "at",
    "nearby",
    "area",

    # Roman Urdu
    "main",
    "mein",
    "me",
    "may",
    "ke",
    "ki",
    "ka",
    "kay",
    "k",
    "ko",
    "par",
    "pe",
    "se",
    "per",
    "dhoondo",
    "dhundo",
    "dhoondho",
    "talash",
    "karo",
    "karain",
    "karein",
    "chahiye",
    "chaheye",
    "chahta",
    "chahti",
    "do",
    "dein",
    "batao",
    "btao",
    "bata",
    "dikhao",
    "dikha",
    "mujhe",
    "mjy",
    "mujhy",
    "please",
    "plz",
    "hai",
    "hain",
    "ho",
    "hoga",
    "tha",
    "thi",
    "thay",
    "nay",
    "koi",
    "jo",
    "wo",
    "woh",
    "yeh",
    "ye",
    "is",
    "us",
    "un",
    "lekin",
    "magar",
    "aur",
    "or",
    "ya",
    "to",
    "bhi",
    "hi",
    "he",
    "kya",
    "kyun",
    "kaise",
    "kahan",
    "kis",
    "kitna",
    "kitne",
    "kitni",
]


# ============================================================
# NORMALIZE QUERY
# ============================================================

def normalize_query(query):

    query = query.lower().strip()

    # Replace punctuation
    query = re.sub(
        r"[?!,.;:/\\|]+",
        " ",
        query
    )

    # Multiple spaces → one
    query = re.sub(
        r"\s+",
        " ",
        query
    )

    return query.strip()


# ============================================================
# FIND BUSINESS TYPE
# ============================================================

def find_business_type(query):

    words = sorted(
        BUSINESS_TYPES.keys(),
        key=len,
        reverse=True
    )

    for word in words:

        pattern = (
            rf"\b{re.escape(word)}\b"
        )

        if re.search(
            pattern,
            query
        ):

            return BUSINESS_TYPES[word], word

    return None, None


# ============================================================
# REMOVE BUSINESS TYPE
# ============================================================

def remove_business_type(
    query,
    business_word
):

    if not business_word:
        return query

    query = re.sub(
        rf"\b{re.escape(business_word)}\b",
        " ",
        query,
        flags=re.IGNORECASE
    )

    return query


# ============================================================
# EXTRACT LOCATION (enhanced)
# ============================================================

def extract_location(
    query,
    business_word
):

    # Remove business type
    location = remove_business_type(
        query,
        business_word
    )

    # ========================================================
    # Remove common natural-language words (phrases first)
    # ========================================================

    phrases = [
        "looking for",
        "looking",
        "find me",
        "find",
        "search for",
        "search",
        "show me",
        "show",
        "give me",
        "give",
        "i need",
        "need",
        "i want",
        "want",
        "mujhe",
        "mujhy",
        "mjy",
        "please",
        "plz",
        "dhoondo",
        "dhundo",
        "dhoondho",
        "talash karo",
        "talash",
        "karo",
        "karain",
        "karein",
        "chahiye",
        "chaheye",
        "batao",
        "btao",
        "dikhao",
        "dikha",
    ]

    for phrase in phrases:

        location = re.sub(
            rf"\b{re.escape(phrase)}\b",
            " ",
            location,
            flags=re.IGNORECASE
        )

    # ========================================================
    # Roman Urdu location words
    # ========================================================

    location_words = [
        "main",
        "mein",
        "me",
        "may",
        "ke",
        "ki",
        "ka",
        "kay",
        "k",
        "ko",
        "par",
        "pe",
        "near",
        "around",
        "in",
        "at",
        "se",
        "per",
        "hai",
        "hain",
        "ho",
        "hoga",
        "tha",
        "thi",
        "thay",
        "nay",
        "koi",
        "jo",
        "wo",
        "woh",
        "yeh",
        "ye",
        "is",
        "us",
        "un",
        "lekin",
        "magar",
        "aur",
        "or",
        "ya",
        "to",
        "bhi",
        "hi",
        "he",
        "kya",
        "kyun",
        "kaise",
        "kahan",
        "kis",
        "kitna",
        "kitne",
        "kitni",
        "please",
        "plz",
    ]

    for word in location_words:

        location = re.sub(
            rf"\b{re.escape(word)}\b",
            " ",
            location,
            flags=re.IGNORECASE
        )

    # ========================================================
    # Clean
    # ========================================================

    location = re.sub(
        r"\s+",
        " ",
        location
    ).strip()

    return location


# ============================================================
# PARSE NATURAL LANGUAGE QUERY
# ============================================================

def parse_query(user_query):

    query = normalize_query(
        user_query
    )

    business_type, business_word = (
        find_business_type(query)
    )

    if not business_type:

        return None

    location = extract_location(
        query,
        business_word
    )

    if not location:

        return None

    return {
        "business_type": business_type,
        "business_word": business_word,
        "location": location
    }


# ============================================================
# GEOCODING
# ============================================================

def geocode(location):

    print(
        f"\n📍 Searching location:"
        f" {location}"
    )

    params = {
        "q": location,
        "format": "json",
        "limit": 1,
        "addressdetails": 1
    }

    response = requests.get(
        NOMINATIM_URL,
        params=params,
        headers=HEADERS,
        timeout=30
    )

    response.raise_for_status()

    results = response.json()

    if not results:

        raise Exception(
            f"Location not found: {location}"
        )

    result = results[0]

    latitude = float(
        result["lat"]
    )

    longitude = float(
        result["lon"]
    )

    print(
        f"✓ Location found:"
        f" {result.get('display_name')}"
    )

    return latitude, longitude


# ============================================================
# SEARCH OPENSTREETMAP
# ============================================================

def search_businesses(
    osm_key,
    osm_value,
    latitude,
    longitude,
    radius=10000
):

    print(
        "\n🔎 Searching OpenStreetMap..."
    )

    query = f"""
    [out:json][timeout:120];

    (
        node["{osm_key}"="{osm_value}"]
        (around:{radius},{latitude},{longitude});

        way["{osm_key}"="{osm_value}"]
        (around:{radius},{latitude},{longitude});

        relation["{osm_key}"="{osm_value}"]
        (around:{radius},{latitude},{longitude});
    );

    out center tags;
    """

    response = requests.post(
        OVERPASS_URL,
        data=query,
        headers=HEADERS,
        timeout=150
    )

    response.raise_for_status()

    data = response.json()

    elements = data.get(
        "elements",
        []
    )

    print(
        f"✓ Found {len(elements)} raw results"
    )

    return elements


# ============================================================
# CLEAN TEXT
# ============================================================

def clean_text(value):

    if value is None:
        return ""

    value = str(value)

    value = re.sub(
        r"\s+",
        " ",
        value
    )

    return value.strip()


# ============================================================
# BUILD ADDRESS
# ============================================================

def build_address(tags):

    parts = []

    fields = [
        "addr:housenumber",
        "addr:street",
        "addr:neighbourhood",
        "addr:suburb",
        "addr:city",
        "addr:postcode"
    ]

    for field in fields:

        value = tags.get(
            field
        )

        if value:

            parts.append(
                clean_text(value)
            )

    return ", ".join(parts)


# ============================================================
# EXTRACT BUSINESS
# ============================================================

def extract_business(element):

    tags = element.get(
        "tags",
        {}
    )

    name = (
        tags.get("name")
        or tags.get("name:en")
        or tags.get("official_name")
    )

    phone = (
        tags.get("phone")
        or tags.get("contact:phone")
        or tags.get("mobile")
    )

    email = (
        tags.get("email")
        or tags.get("contact:email")
    )

    website = (
        tags.get("website")
        or tags.get("contact:website")
        or tags.get("url")
    )

    return {

        "Name": clean_text(
            name
        ),

        "Phone": clean_text(
            phone
        ),

        "Email": clean_text(
            email
        ),

        "Website": clean_text(
            website
        ),

        "Address": build_address(
            tags
        ),

        "City": clean_text(
            tags.get("addr:city")
        ),

        "Postal Code": clean_text(
            tags.get("addr:postcode")
        ),

        "Opening Hours": clean_text(
            tags.get("opening_hours")
        ),

        "Cuisine": clean_text(
            tags.get("cuisine")
        ),

        "Rating": clean_text(
            tags.get("stars")
        )
    }


# ============================================================
# CLEAN DATA
# ============================================================

def clean_data(data):

    cleaned = []

    seen = set()

    for business in data:

        name = business.get(
            "Name",
            ""
        ).strip()

        if not name:
            continue

        key = (
            name.lower(),
            business.get(
                "Address",
                ""
            ).lower()
        )

        if key in seen:
            continue

        seen.add(
            key
        )

        cleaned.append(
            business
        )

    return cleaned


# ============================================================
# CREATE DATAFRAME (sort with phone numbers first)
# ============================================================

def create_dataframe(data):

    columns = [
        "Name",
        "Phone",
        "Email",
        "Website",
        "Address",
        "City",
        "Postal Code",
        "Opening Hours",
        "Cuisine",
        "Rating"
    ]

    df = pd.DataFrame(
        data
    )

    # Ensure columns exist
    for column in columns:

        if column not in df.columns:

            df[column] = ""

    # Keep only desired columns
    df = df[
        columns
    ]

    # Replace NaN
    df = df.fillna("")

    # Clean strings
    for column in columns:

        df[column] = (
            df[column]
            .astype(str)
            .str.strip()
        )

    # Remove empty names
    df = df[
        df["Name"] != ""
    ]

    # Remove duplicates
    df = df.drop_duplicates(
        subset=[
            "Name",
            "Address"
        ]
    )

    # ---- SORT: Phone numbers first, then by Name ----
    df['_has_phone'] = df['Phone'] != ""
    df = df.sort_values(
        by=['_has_phone', 'Name'],
        ascending=[False, True],
        key=lambda x: x.str.lower() if x.name == 'Name' else x
    )
    df = df.drop(columns=['_has_phone'])

    df = df.reset_index(
        drop=True
    )

    return df


# ============================================================
# EXPORT CSV (with custom base name)
# ============================================================

def export_csv(df, base_name="businesses"):

    filename = f"{base_name}.csv"

    df.to_csv(
        filename,
        index=False,
        encoding="utf-8-sig"
    )

    print(
        f"\n✓ CSV created:"
        f" {filename}"
    )

    print(
        f"✓ Records:"
        f" {len(df)}"
    )


# ============================================================
# EXPORT EXCEL (with custom base name)
# ============================================================

def export_excel(df, base_name="businesses"):

    filename = f"{base_name}.xlsx"

    df.to_excel(
        filename,
        index=False,
        sheet_name="Businesses",
        engine="openpyxl"
    )

    workbook = load_workbook(
        filename
    )

    worksheet = workbook[
        "Businesses"
    ]

    # Freeze header
    worksheet.freeze_panes = "A2"

    # Header styling
    for cell in worksheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F2937"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    worksheet.row_dimensions[
        1
    ].height = 30

    # Column widths
    widths = {
        "A": 32,
        "B": 22,
        "C": 32,
        "D": 42,
        "E": 55,
        "F": 22,
        "G": 15,
        "H": 30,
        "I": 25,
        "J": 12
    }

    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # Add filter
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # Add Excel table
    if worksheet.max_row >= 2:

        last_column = (
            worksheet.max_column
        )

        last_column_letter = (
            worksheet.cell(
                row=1,
                column=last_column
            ).column_letter
        )

        table = Table(
            displayName="BusinessData",
            ref=(
                f"A1:"
                f"{last_column_letter}"
                f"{worksheet.max_row}"
            )
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        worksheet.add_table(
            table
        )

    # Website hyperlinks
    website_column = 4

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        cell = worksheet.cell(
            row=row,
            column=website_column
        )

        website = cell.value

        if website:

            website = str(
                website
            ).strip()

            if not website.startswith(
                (
                    "http://",
                    "https://"
                )
            ):

                website = (
                    "https://"
                    + website
                )

            cell.hyperlink = website
            cell.style = "Hyperlink"

    # Wrap text
    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    workbook.save(
        filename
    )

    print(
        f"✓ Excel created:"
        f" {filename}"
    )

    print(
        f"✓ Records:"
        f" {len(df)}"
    )


# ============================================================
# PDF EXPORT (pretty, with header background & alternating rows)
# ============================================================

def sanitize_text(text):
    """Replace non-ASCII characters with '?' for core font."""
    if not text:
        return ""
    return ''.join(c if ord(c) < 128 else '?' for c in str(text))


def export_pdf(df, base_name="businesses"):
    """
    Generate a beautiful PDF table with:
    - Colored header (dark blue)
    - Alternating row colors (light grey / white)
    - Proper column widths (Phone column enlarged)
    - Wrapped text
    - Record count footer
    """
    if df.empty:
        print("⚠️ No data to export to PDF.")
        return

    filename = f"{base_name}.pdf"
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    font_name = "Helvetica"

    # ---- Title ----
    pdf.set_font(font_name, 'B', 14)
    title = f"Business List - {base_name.replace('_', ' ').title()}"
    title = sanitize_text(title)
    pdf.cell(0, 10, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(6)

    # ---- Column definitions ----
    headers = list(df.columns)
    # Adjusted widths: Name 36, Phone 24 (wider), Email 22, Website 28, Address 40, City 18, Postal 16, Hours 20, Cuisine 18, Rating 12
    col_widths = [36, 24, 22, 28, 40, 18, 16, 20, 18, 12]  # sum = 234, fits nicely

    # ---- Calculate table start X to center ----
    total_width = sum(col_widths)
    page_width = pdf.w - 2 * pdf.l_margin
    start_x = (page_width - total_width) / 2 + pdf.l_margin
    pdf.set_x(start_x)

    # ---- Header row (with background) ----
    pdf.set_font(font_name, 'B', 9)
    # Save fill color and draw filled cells
    pdf.set_fill_color(30, 40, 60)  # dark blue
    pdf.set_text_color(255, 255, 255)  # white
    for i, header in enumerate(headers):
        h = sanitize_text(header)
        pdf.cell(col_widths[i], 8, h, border=1, align='C', fill=True)
    pdf.ln()
    # Reset text color
    pdf.set_text_color(0, 0, 0)

    # ---- Data rows (alternating colors) ----
    pdf.set_font(font_name, '', 8)
    line_height = 5  # mm per line

    for row_idx, (_, row) in enumerate(df.iterrows()):
        # Prepare each column's text (sanitized)
        row_texts = []
        max_lines = 1
        for i, col in enumerate(headers):
            text = str(row[col]) if pd.notna(row[col]) else ""
            text = sanitize_text(text)
            row_texts.append(text)

        # Determine number of lines needed per column (wrap by spaces)
        lines_per_col = []
        for i, text in enumerate(row_texts):
            words = text.split(' ')
            lines = []
            current_line = ""
            for word in words:
                if pdf.get_string_width(current_line + " " + word) <= col_widths[i] - 2:
                    if current_line:
                        current_line += " " + word
                    else:
                        current_line = word
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word
            if current_line:
                lines.append(current_line)
            if not lines:
                lines = [""]
            lines_per_col.append(lines)
            if len(lines) > max_lines:
                max_lines = len(lines)

        # Calculate row height
        row_height = max(8, max_lines * line_height + 2)

        # Check page break
        if pdf.get_y() + row_height > pdf.h - pdf.b_margin - 10:  # leave space for footer
            pdf.add_page()
            # Re-print header on new page
            pdf.set_font(font_name, 'B', 9)
            pdf.set_fill_color(30, 40, 60)
            pdf.set_text_color(255, 255, 255)
            pdf.set_x(start_x)
            for i, header in enumerate(headers):
                h = sanitize_text(header)
                pdf.cell(col_widths[i], 8, h, border=1, align='C', fill=True)
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
            pdf.set_font(font_name, '', 8)

        # Choose background color for this row
        if row_idx % 2 == 0:
            fill = True
            pdf.set_fill_color(240, 240, 240)  # light grey
        else:
            fill = False
            pdf.set_fill_color(255, 255, 255)  # white (but fill=False so it's transparent)

        # Print the row using multi_cell (fill only for even rows)
        y_start = pdf.get_y()
        for i, lines in enumerate(lines_per_col):
            pdf.set_xy(start_x + sum(col_widths[:i]), y_start)
            # Use multi_cell with fill if even row
            pdf.multi_cell(col_widths[i], line_height, "\n".join(lines),
                           border=1, align='L', new_x=XPos.RIGHT, new_y=YPos.NEXT,
                           fill=fill)
        pdf.set_y(y_start + row_height)

    # ---- Footer: total records ----
    pdf.set_y(pdf.h - pdf.b_margin - 6)
    pdf.set_font(font_name, 'I', 9)
    pdf.cell(0, 6, f"Total records: {len(df)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

    # ---- Save PDF ----
    pdf.output(filename)
    print(f"✓ PDF created: {filename}")
    print(f"✓ Records: {len(df)}")


# ============================================================
# DISPLAY RESULTS
# ============================================================

def display_results(df):

    if df.empty:

        print(
            "\nNo businesses found."
        )

        return

    print(
        "\n"
        + "=" * 70
    )

    print(
        f"FOUND {len(df)} BUSINESSES"
    )

    print(
        "=" * 70
    )

    for index, row in df.iterrows():

        print(
            f"\n{index + 1}. "
            f"{row['Name']}"
        )

        print(
            f"   📞 Phone: "
            f"{row['Phone'] or 'N/A'}"
        )

        print(
            f"   ✉ Email: "
            f"{row['Email'] or 'N/A'}"
        )

        print(
            f"   🌐 Website: "
            f"{row['Website'] or 'N/A'}"
        )

        print(
            f"   📍 Address: "
            f"{row['Address'] or 'N/A'}"
        )

        print(
            f"   🏙 City: "
            f"{row['City'] or 'N/A'}"
        )


# ============================================================
# GENERATE UNIQUE FILENAME
# ============================================================

def generate_filename(parsed):
    business_word = parsed['business_word']
    location = parsed['location']
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', f"{business_word}_{location}".lower())
    sanitized = re.sub(r'_+', '_', sanitized)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"businesses_{sanitized}_{timestamp}"


# ============================================================
# SCRAPER
# ============================================================

def scrape(user_query):

    print(
        "\n"
        + "=" * 70
    )

    print(
        "FREE NATURAL LANGUAGE BUSINESS SCRAPER"
    )

    print(
        "=" * 70
    )

    # Parse query
    parsed = parse_query(
        user_query
    )

    if not parsed:

        print(
            "\n❌ I couldn't understand that query."
        )

        print(
            "\nTry:"
        )

        print(
            "  restaurants in I-8 Islamabad"
        )

        print(
            "  i8 Islamabad main restaurants dhoondo"
        )

        print(
            "  Islamabad mein hotels find karo"
        )

        print(
            "  Rawalpindi main cafes chahiye"
        )

        return pd.DataFrame(), None

    osm_key, osm_value = (
        parsed["business_type"]
    )

    location = parsed[
        "location"
    ]

    print(
        f"\n🏢 Type:"
        f" {parsed['business_word']}"
    )

    print(
        f"📍 Location:"
        f" {location}"
    )

    # Respect Nominatim
    time.sleep(1)

    latitude, longitude = geocode(
        location
    )

    elements = search_businesses(
        osm_key,
        osm_value,
        latitude,
        longitude
    )

    if not elements:

        return pd.DataFrame(), parsed

    print(
        "\n📦 Extracting business data..."
    )

    businesses = []

    for index, element in enumerate(
        elements,
        start=1
    ):

        try:

            business = extract_business(
                element
            )

            businesses.append(
                business
            )

        except Exception as error:

            print(
                f"Error on result "
                f"{index}: {error}"
            )

    businesses = clean_data(
        businesses
    )

    df = create_dataframe(
        businesses
    )

    return df, parsed


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        """
╔══════════════════════════════════════════════════════════════╗
║             FREE BUSINESS DATA SCRAPER                      ║
╠══════════════════════════════════════════════════════════════╣
║                                                              ║
║ English:                                                    ║
║ restaurants in I-8 Islamabad                                ║
║ hotels in Islamabad                                        ║
║ cafes in Rawalpindi                                        ║
║                                                              ║
║ Roman Urdu:                                                 ║
║ i8 Islamabad main restaurants dhoondo                       ║
║ Islamabad mein hotels find karo                            ║
║ Rawalpindi main cafes chahiye                              ║
║ Lahore mein schools dhoondo                                ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
"""
    )

    query = input(
        "\n🔎 What are you looking for?\n> "
    ).strip()

    if not query:

        print(
            "\nPlease enter something."
        )

        return

    try:

        df, parsed = scrape(
            query
        )

        if df.empty:

            print(
                "\n❌ No data to export."
            )

            return

        # Generate unique base filename
        base_name = generate_filename(parsed)

        # Display
        display_results(
            df
        )

        # Export
        print(
            "\n📤 Exporting..."
        )

        export_csv(
            df,
            base_name
        )

        export_excel(
            df,
            base_name
        )

        export_pdf(
            df,
            base_name
        )

        print(
            "\n"
            + "=" * 70
        )

        print(
            "✅ COMPLETE"
        )

        print(
            "=" * 70
        )

        print(
            f"\nTotal businesses:"
            f" {len(df)}"
        )

        print(
            "\nFiles created:"
        )

        print(
            f"  📄 {base_name}.csv"
        )

        print(
            f"  📊 {base_name}.xlsx"
        )

        print(f"  📕 {base_name}.pdf")

    except requests.exceptions.Timeout:

        print(
            "\n⏱ Request timed out."
        )

        print(
            "Please try again."
        )

    except requests.exceptions.RequestException as error:

        print(
            f"\n🌐 Network error:"
            f" {error}"
        )

    except Exception as error:

        print(
            f"\n❌ Error:"
            f" {error}"
        )


if __name__ == "__main__":

    main()